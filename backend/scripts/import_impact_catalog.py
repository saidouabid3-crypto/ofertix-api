from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator
from urllib.parse import parse_qs, unquote, urlparse

BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_FEEDS_DIR = BASE_DIR / "data" / "impact_feeds"
DEFAULT_SINGLE_FEED = BASE_DIR / "data" / "impact_dhgate.txt"
DEFAULT_REPORT_DIR = BASE_DIR / "data" / "import_reports"
DEFAULT_CHECKPOINT_DIR = BASE_DIR / "data" / "import_checkpoints"
SUPPORTED_TEXT_SUFFIXES = {".csv", ".tsv", ".txt"}
SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xls"}
SUPPORTED_SUFFIXES = SUPPORTED_TEXT_SUFFIXES | SUPPORTED_EXCEL_SUFFIXES
DEFAULT_COUNTRIES = ["es", "ma", "dz", "fr", "pt", "it", "de", "uk", "us", "ca", "eg", "sa", "ae", "mx"]
CSV_ENCODINGS = ("utf-8-sig", "utf-16", "latin-1")
UNAVAILABLE_STOCK = {
    "outofstock",
    "out of stock",
    "soldout",
    "sold out",
    "unavailable",
    "discontinued",
    "false",
    "0",
}

try:
    from core.elite_categories import classify_elite_category, map_to_elite_category
    from utils.product_standard import product_fingerprint, score_product
except ModuleNotFoundError:
    sys.path.insert(0, str(BASE_DIR))
    from core.elite_categories import classify_elite_category, map_to_elite_category
    from utils.product_standard import product_fingerprint, score_product


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(value))).strip()


def normalize_price(value: Any) -> float | None:
    raw = clean_text(value)
    if not raw:
        return None
    raw = re.sub(r"[^0-9,.\-]", "", raw)
    if not raw or raw in {"-", ".", ","}:
        return None

    comma = raw.rfind(",")
    dot = raw.rfind(".")
    if comma >= 0 and dot >= 0:
        if comma > dot:
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif comma >= 0:
        whole, _, frac = raw.partition(",")
        raw = whole + frac if len(frac) == 3 and len(whole) > 1 else raw.replace(",", ".")
    elif dot >= 0:
        whole, _, frac = raw.partition(".")
        if len(frac) == 3 and len(whole) > 1:
            raw = whole + frac

    try:
        price = float(raw)
    except ValueError:
        return None
    return round(price, 2) if price > 0 else None


def normalize_int(value: Any) -> int | None:
    price = normalize_price(value)
    if price is None:
        return None
    return int(round(price))


def normalize_currency(value: Any) -> str:
    currency = clean_text(value).upper()
    if not currency:
        return ""
    currency = re.sub(r"[^A-Z]", "", currency)
    return currency[:3]


def first_present(row: dict[str, Any], *names: str) -> str:
    for name in names:
        value = clean_text(row.get(name))
        if value:
            return value
    return ""


def extract_original_url(affiliate_url: str) -> str:
    affiliate_url = clean_text(affiliate_url)
    if not affiliate_url:
        return ""
    try:
        query = parse_qs(urlparse(affiliate_url).query)
        if query.get("u"):
            return unquote(query["u"][0])
    except Exception:
        return ""
    return ""


def first_valid_image(primary: str, additional: str = "") -> tuple[str, list[str]]:
    images: list[str] = []
    for candidate in [primary, *re.split(r"[\s,;|]+", additional or "")]:
        image = clean_text(candidate)
        if image.startswith(("http://", "https://")) and image not in images:
            images.append(image)
    return (images[0], images[:3]) if images else ("", [])


def short_name(title: str) -> str:
    title = re.sub(r"\s+", " ", title).strip()
    title = re.sub(
        r"(?i)\b(wholesale|dropshipping|free shipping|hot sale|new arrival|best price)\b",
        "",
        title,
    ).strip()
    first = re.split(r"[,|\-]", title)[0].strip()
    name = first if len(first) >= 10 else title
    if len(name) > 70:
        name = name[:70].rsplit(" ", 1)[0]
    return name or title[:70] or "Product"


def stock_is_available(value: Any) -> bool:
    stock = clean_text(value).lower().replace("_", " ")
    compact = stock.replace(" ", "")
    if not stock:
        return True
    return stock not in UNAVAILABLE_STOCK and compact not in UNAVAILABLE_STOCK


def availability_status(value: Any) -> str:
    stock = clean_text(value)
    if not stock:
        return "unknown"
    return "in_stock" if stock_is_available(stock) else "out_of_stock"


def stable_impact_doc_id(program_id: str, source_id: str, affiliate_url: str = "") -> str:
    program = re.sub(r"[^A-Za-z0-9_-]+", "_", clean_text(program_id) or "unknown").strip("_")
    source = clean_text(source_id)
    safe_source = re.sub(r"[^A-Za-z0-9_-]+", "_", source).strip("_")
    if safe_source and len(safe_source) <= 90:
        return f"impact_{program}_{safe_source}"
    digest = hashlib.sha1(f"{program_id}|{source}|{affiliate_url}".encode("utf-8")).hexdigest()[:24]
    return f"impact_{program}_{digest}"


@dataclass
class CatalogInspection:
    path: str
    size: int
    encoding: str
    delimiter: str
    format: str
    headers: list[str]
    row_count: int
    first_valid_row: dict[str, str] = field(default_factory=dict)
    last_valid_row: dict[str, str] = field(default_factory=dict)


@dataclass
class ImportStats:
    scanned: int = 0
    valid: int = 0
    upserted: int = 0
    imported: int = 0
    updated: int = 0
    skipped: int = 0
    invalid: int = 0
    failed: int = 0
    duplicates_in_run: int = 0
    batches_committed: int = 0
    skip_reasons: dict[str, int] = field(default_factory=dict)
    failures: list[dict[str, Any]] = field(default_factory=list)

    def skip(self, reason: str) -> None:
        self.skipped += 1
        self.skip_reasons[reason] = self.skip_reasons.get(reason, 0) + 1

    def bad(self, reason: str) -> None:
        self.invalid += 1
        self.skip(reason)


def detect_encoding(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in CSV_ENCODINGS:
        try:
            raw.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8"


def detect_delimiter(sample: str, suffix: str) -> str:
    if suffix == ".tsv":
        return "\t"
    try:
        return csv.Sniffer().sniff(sample[:8192], delimiters="\t,;|").delimiter
    except csv.Error:
        counts = {delimiter: sample.count(delimiter) for delimiter in ("\t", ",", ";", "|")}
        return max(counts, key=counts.get)


def iter_text_rows(path: Path, *, encoding: str | None = None, delimiter: str | None = None) -> Iterator[dict[str, Any]]:
    encoding = encoding or detect_encoding(path)
    with path.open("r", encoding=encoding, errors="replace", newline="") as file:
        sample = file.read(8192)
        file.seek(0)
        delimiter = delimiter or detect_delimiter(sample, path.suffix.lower())
        reader = csv.DictReader(file, delimiter=delimiter)
        for row in reader:
            yield dict(row)


def iter_excel_rows(path: Path) -> Iterator[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to import Excel catalog files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(cell) for cell in next(rows, [])]
    for row in rows:
        yield {headers[index]: value for index, value in enumerate(row) if index < len(headers)}


def iter_catalog_rows(path: Path) -> Iterator[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in SUPPORTED_TEXT_SUFFIXES:
        yield from iter_text_rows(path)
        return
    if suffix in SUPPORTED_EXCEL_SUFFIXES:
        yield from iter_excel_rows(path)
        return
    raise ValueError(f"Unsupported catalog format: {path}")


def read_headers(path: Path) -> tuple[str, str, list[str]]:
    suffix = path.suffix.lower()
    if suffix in SUPPORTED_EXCEL_SUFFIXES:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            return "binary", "", []
        workbook = load_workbook(path, read_only=True, data_only=True)
        row = next(workbook.active.iter_rows(values_only=True), [])
        return "binary", "", [clean_text(cell) for cell in row]

    encoding = detect_encoding(path)
    with path.open("r", encoding=encoding, errors="replace", newline="") as file:
        sample = file.read(8192)
        delimiter = detect_delimiter(sample, suffix)
        file.seek(0)
        reader = csv.reader(file, delimiter=delimiter)
        headers = next(reader, [])
    return encoding, delimiter, [clean_text(header) for header in headers]


def summarize_row(row: dict[str, Any]) -> dict[str, str]:
    keys = (
        "Sku",
        "Program Id",
        "Program Names",
        "Name",
        "Url",
        "Original Url",
        "Image Url",
        "Current Price",
        "Original Price",
        "Currency",
        "Category Name",
        "Category Path",
        "Stock Availability",
        "Last Updated",
    )
    return {key: clean_text(row.get(key))[:140] for key in keys if clean_text(row.get(key))}


def inspect_catalog_file(path: Path) -> CatalogInspection:
    encoding, delimiter, headers = read_headers(path)
    row_count = 0
    first: dict[str, str] = {}
    last: dict[str, str] = {}
    for row in iter_catalog_rows(path):
        if not any(clean_text(value) for value in row.values()):
            continue
        row_count += 1
        summary = summarize_row(row)
        if not first:
            first = summary
        last = summary

    suffix = path.suffix.lower()
    file_format = "XLSX" if suffix == ".xlsx" else "XLS" if suffix == ".xls" else "TSV" if delimiter == "\t" else "CSV"
    return CatalogInspection(
        path=str(path),
        size=path.stat().st_size,
        encoding=encoding,
        delimiter="TAB" if delimiter == "\t" else delimiter,
        format=file_format,
        headers=headers,
        row_count=row_count,
        first_valid_row=first,
        last_valid_row=last,
    )


def discover_catalog_files(paths: Iterable[str] | None, folder: str | None) -> list[Path]:
    candidates: list[Path] = []
    if paths:
        candidates.extend(Path(path) for path in paths)
    else:
        if DEFAULT_SINGLE_FEED.exists():
            candidates.append(DEFAULT_SINGLE_FEED)
        feeds_dir = Path(folder) if folder else DEFAULT_FEEDS_DIR
        if feeds_dir.exists():
            candidates.extend(sorted(path for path in feeds_dir.iterdir() if path.is_file()))

    result: list[Path] = []
    seen: set[str] = set()
    for path in candidates:
        resolved = path.resolve()
        if resolved.name.upper().startswith("README"):
            continue
        if resolved.suffix.lower() not in SUPPORTED_SUFFIXES:
            continue
        key = str(resolved).lower()
        if key not in seen:
            result.append(resolved)
            seen.add(key)
    return result


def build_impact_product(row: dict[str, Any]) -> tuple[str | None, dict[str, Any] | None, str | None]:
    sku = first_present(row, "Sku", "SKU", "Product Id", "Product ID")
    program_id = first_present(row, "Program Id", "Program ID", "ProgramId")
    source_id = sku or first_present(row, "Gtin", "GTIN", "Catalog Id")
    program_name = first_present(row, "Program Names", "Program Name", "Merchant", "Store")
    title = first_present(row, "Name", "Title", "Product Name")
    affiliate_url = first_present(row, "Url", "URL", "Affiliate Url", "Affiliate URL")
    original_url = first_present(row, "Original Url", "Original URL", "Product Url", "Product URL")
    image, images = first_valid_image(
        first_present(row, "Image Url", "Image URL", "Image"),
        first_present(row, "Additional ImageUrls", "Additional Image URLs", "ImageUrls"),
    )
    current_price = normalize_price(first_present(row, "Current Price", "Price", "Sale Price"))
    original_price = normalize_price(first_present(row, "Original Price", "Was Price", "List Price"))
    currency = normalize_currency(first_present(row, "Currency", "Price Currency"))
    stock = first_present(row, "Stock Availability", "Availability")

    if not (source_id or (program_id and affiliate_url)):
        return None, None, "missing_stable_id"
    if not title:
        return None, None, "missing_name"
    if not affiliate_url.startswith(("http://", "https://")):
        return None, None, "missing_or_invalid_url"
    if not image:
        return None, None, "missing_image"
    if current_price is None:
        return None, None, "missing_or_invalid_current_price"
    if not currency:
        return None, None, "missing_currency"
    if not stock_is_available(stock):
        return None, None, "out_of_stock"

    product_url = original_url if original_url.startswith(("http://", "https://")) else extract_original_url(affiliate_url)
    category_name = first_present(row, "Category Name", "Category")
    category_path = first_present(row, "Category Path")
    category_source = category_path or category_name
    category, confidence, reason = classify_elite_category(
        {
            "name": title,
            "title": title,
            "description": first_present(row, "Description", "Bullets"),
            "category": category_source,
            "sourceCategory": category_name,
            "storeCategory": category_path,
        }
    )
    category = map_to_elite_category(category)

    discount = normalize_int(first_present(row, "Discount Percentage", "Discount"))
    if (not discount or discount <= 0) and original_price and original_price > current_price:
        discount = int(round(((original_price - current_price) / original_price) * 100))
    if not discount or discount <= 0 or discount > 95:
        discount = 0

    doc_id = stable_impact_doc_id(program_id, source_id or affiliate_url, affiliate_url)
    now = utc_now()
    product: dict[str, Any] = {
        "id": doc_id,
        "sku": sku,
        "programId": program_id,
        "programName": program_name,
        "catalogId": first_present(row, "Catalog Id"),
        "name": short_name(title),
        "fullTitle": title,
        "description": first_present(row, "Description", "Bullets") or title,
        "image": image,
        "mainImage": image,
        "images": images,
        "newPrice": current_price,
        "currency": currency,
        "store": program_name or "Impact",
        "source": "impact",
        "sourceType": "affiliate_product",
        "affiliateNetwork": "impact",
        "affiliateUrl": affiliate_url,
        "productUrl": product_url,
        "sourceCategoryName": category_name,
        "sourceCategoryPath": category_path,
        "category": category,
        "categoryGroup": category,
        "categoryConfidence": confidence,
        "categoryReason": reason,
        "stockAvailability": stock,
        "availabilityStatus": availability_status(stock),
        "condition": first_present(row, "Condition"),
        "promotions": first_present(row, "Promotions"),
        "lastUpdatedFromFeed": first_present(row, "Last Updated"),
        "countryCode": "global",
        "country": "global",
        "isExplicitlyGlobal": True,
        "availableCountries": DEFAULT_COUNTRIES,
        "shipsTo": DEFAULT_COUNTRIES,
        "marketType": "international",
        "pickupOnly": False,
        "status": "active",
        "visibleToUsers": True,
        "publicVisible": True,
        "isExpired": False,
        "isAffiliateProduct": True,
        "itemType": "product",
        "listingType": "affiliate_product",
        "contactSellerEnabled": False,
        "priceAccuracy": "feed",
        "priceSource": "impact_catalog",
        "finalPriceInStore": True,
        "importSource": "impact_catalog",
        "importedAt": now,
        "lastSyncedAt": now,
        "updatedAt": now,
    }
    if original_price and original_price > current_price:
        product["oldPrice"] = original_price
    if discount:
        product["discount"] = discount

    deal, trust, quality, verdict = score_product(product)
    product.update(
        {
            "dealScore": deal,
            "trustScore": trust,
            "qualityScore": quality,
            "verdict": verdict,
            "fingerprint": product_fingerprint(product),
        }
    )
    return doc_id, product, None


class DryRunSink:
    def __init__(self) -> None:
        self.existing: set[str] = set()

    def upsert(self, doc_id: str, product: dict[str, Any]) -> str:
        status = "updated" if doc_id in self.existing else "imported"
        self.existing.add(doc_id)
        return status

    def commit(self) -> None:
        return None


class FirestoreSink:
    def __init__(self, *, batch_size: int) -> None:
        from core.firebase import db

        if db is None:
            raise RuntimeError("Firebase is not configured")
        self.db = db
        self.batch_size = batch_size
        self.pending: list[tuple[str, dict[str, Any]]] = []
        self.commits = 0
        self.imported = 0
        self.updated = 0

    def upsert(self, doc_id: str, product: dict[str, Any]) -> str:
        self.pending.append((doc_id, product))
        if len(self.pending) >= self.batch_size:
            self.commit()
        return "queued"

    def commit(self) -> None:
        if not self.pending:
            return

        collection = self.db.collection("products")
        refs = [collection.document(doc_id) for doc_id, _ in self.pending]
        existing_ids = {doc.id for doc in self.db.get_all(refs) if doc.exists}

        batch = self.db.batch()
        for doc_id, product in self.pending:
            batch.set(collection.document(doc_id), product, merge=True)
            if doc_id in existing_ids:
                self.updated += 1
            else:
                self.imported += 1
        batch.commit()
        self.pending = []
        self.commits += 1


def load_checkpoint(path: Path) -> dict[str, int]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return {str(k): int(v) for k, v in data.get("files", {}).items()}


def save_checkpoint(path: Path, checkpoints: dict[str, int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"updatedAt": utc_now(), "files": checkpoints}
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def import_catalog_files(
    files: list[Path],
    *,
    apply: bool,
    limit: int | None,
    batch_size: int,
    checkpoint_path: Path | None = None,
    resume: bool = False,
    progress_every: int = 5000,
) -> dict[str, Any]:
    stats = ImportStats()
    sink: Any = FirestoreSink(batch_size=batch_size) if apply else DryRunSink()
    seen_doc_ids: set[str] = set()
    checkpoints = load_checkpoint(checkpoint_path) if checkpoint_path and resume else {}
    file_results: list[dict[str, Any]] = []

    for file_path in files:
        file_key = str(file_path.resolve())
        resume_after = checkpoints.get(file_key, 0)
        file_scanned = 0
        file_upserted = 0
        file_skipped = 0
        for row_number, row in enumerate(iter_catalog_rows(file_path), start=1):
            if resume_after and row_number <= resume_after:
                continue
            if limit is not None and stats.upserted >= limit:
                break

            stats.scanned += 1
            file_scanned += 1
            doc_id, product, reason = build_impact_product(row)
            if reason:
                stats.bad(reason)
                file_skipped += 1
                continue
            assert doc_id is not None and product is not None

            if doc_id in seen_doc_ids:
                stats.duplicates_in_run += 1
                stats.skip("duplicate_in_input")
                file_skipped += 1
                continue
            seen_doc_ids.add(doc_id)
            stats.valid += 1

            try:
                result = sink.upsert(doc_id, product)
                stats.upserted += 1
                file_upserted += 1
                if result == "updated":
                    stats.updated += 1
                elif result == "imported":
                    stats.imported += 1
            except Exception as exc:
                stats.failed += 1
                stats.failures.append({"file": str(file_path), "row": row_number, "error": str(exc)[:300]})

            if apply and checkpoint_path and stats.upserted and stats.upserted % batch_size == 0:
                checkpoints[file_key] = row_number
                save_checkpoint(checkpoint_path, checkpoints)
            if progress_every and stats.scanned % progress_every == 0:
                print(
                    "Progress: "
                    f"scanned={stats.scanned} upserted={stats.upserted} "
                    f"skipped={stats.skipped} failed={stats.failed}",
                    flush=True,
                )

        file_results.append(
            {
                "file": str(file_path),
                "scanned": file_scanned,
                "upserted": file_upserted,
                "skipped": file_skipped,
            }
        )
        if apply and checkpoint_path:
            checkpoints[file_key] = resume_after + file_scanned
            save_checkpoint(checkpoint_path, checkpoints)
        if limit is not None and stats.upserted >= limit:
            break

    sink.commit()
    stats.batches_committed = getattr(sink, "commits", 0)
    if apply:
        stats.imported = getattr(sink, "imported", stats.imported)
        stats.updated = getattr(sink, "updated", stats.updated)
    return {
        "mode": "apply" if apply else "dry-run",
        "collection": "products",
        "generatedAt": utc_now(),
        "stats": stats.__dict__,
        "files": file_results,
    }


def write_report(report_path: Path, payload: dict[str, Any]) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def default_report_path(mode: str) -> Path:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return DEFAULT_REPORT_DIR / f"impact_catalog_{mode}_{stamp}.json"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import Impact product catalogs into Ofertix Firestore products")
    parser.add_argument("--path", action="append", help="Catalog file path. Repeat for multiple files.")
    parser.add_argument("--folder", default=str(DEFAULT_FEEDS_DIR), help="Folder containing Impact catalog files.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate without writing to Firestore.")
    parser.add_argument("--apply", action="store_true", help="Write valid products to Firestore products.")
    parser.add_argument("--inspect-only", action="store_true", help="Only inspect catalog files.")
    parser.add_argument("--limit", type=int, default=None, help="Maximum valid products to upsert across all files.")
    parser.add_argument("--batch-size", type=int, default=400, help="Firestore batch size, capped below 500.")
    parser.add_argument("--resume", action="store_true", help="Resume from checkpoint.")
    parser.add_argument("--checkpoint", default=str(DEFAULT_CHECKPOINT_DIR / "impact_catalog_checkpoint.json"))
    parser.add_argument("--report", default=None, help="Report JSON output path.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.apply == args.dry_run and not args.inspect_only:
        print("Choose exactly one of --dry-run or --apply, unless using --inspect-only.", file=sys.stderr)
        return 2

    files = discover_catalog_files(args.path, args.folder)
    if not files:
        print("No Impact catalog files found.", file=sys.stderr)
        return 1

    inspections = [inspect_catalog_file(path) for path in files]
    mode = "inspect" if args.inspect_only else "apply" if args.apply else "dry_run"
    report_path = Path(args.report) if args.report else default_report_path(mode)
    payload: dict[str, Any] = {
        "generatedAt": utc_now(),
        "catalogFiles": [inspection.__dict__ for inspection in inspections],
    }

    print(f"Found {len(files)} catalog files")
    for inspection in inspections:
        print(
            f"- {inspection.path} | {inspection.format} | {inspection.encoding} | "
            f"{inspection.delimiter} | rows={inspection.row_count} | size={inspection.size}"
        )

    if not args.inspect_only:
        batch_size = max(1, min(args.batch_size, 450))
        import_result = import_catalog_files(
            files,
            apply=args.apply,
            limit=args.limit,
            batch_size=batch_size,
            checkpoint_path=Path(args.checkpoint),
            resume=args.resume,
        )
        payload["importResult"] = import_result
        stats = import_result["stats"]
        print(
            "Import result: "
            f"scanned={stats['scanned']} upserted={stats['upserted']} "
            f"imported={stats['imported']} updated={stats['updated']} "
            f"skipped={stats['skipped']} invalid={stats['invalid']} failed={stats['failed']}"
        )

    write_report(report_path, payload)
    print(f"Report written: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
